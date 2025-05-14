import sys
import os

from flask import Flask, render_template, request, send_file, jsonify
import PyPDF2
import re
import io
from werkzeug.utils import secure_filename
import logging
import openpyxl
from odf.opendocument import load as load_ods
from odf.table import Table, TableRow, TableCell
from odf.text import P

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 32 * 1024 * 1024  # 32MB

# Configuração do logger do Flask
if not app.debug: # Em produção
    gunicorn_logger = logging.getLogger("gunicorn.error")
    if gunicorn_logger.handlers:
        app.logger.handlers = gunicorn_logger.handlers
        app.logger.setLevel(gunicorn_logger.level)
    else:
        app.logger.setLevel(logging.INFO)
        # Adicionar um handler básico se nenhum estiver configurado (ex: rodando sem Gunicorn)
        if not app.logger.handlers:
            stream_handler = logging.StreamHandler(sys.stdout)
            stream_handler.setFormatter(logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s"))
            app.logger.addHandler(stream_handler)
elif app.debug: # Em debug local
    app.logger.setLevel(logging.DEBUG)
    # Garantir que o handler padrão do Flask (ou um novo) esteja configurado para debug
    if not app.logger.handlers or not any(isinstance(h, logging.StreamHandler) for h in app.logger.handlers):
        app.logger.handlers.clear() # Limpar para evitar duplicatas no reload
        stream_handler = logging.StreamHandler(sys.stdout)
        stream_handler.setFormatter(logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s"))
        app.logger.addHandler(stream_handler)

def extrair_dados_pdf_refatorado(file_stream):
    try:
        pdf_buffer = io.BytesIO()
        if hasattr(file_stream, "save") and callable(getattr(file_stream, "save")):
            file_stream.save(pdf_buffer)
            if hasattr(file_stream, "seek") and callable(getattr(file_stream, "seek")):
                file_stream.seek(0)
        else:
            pdf_buffer.write(file_stream.read())
            if hasattr(file_stream, "seek") and callable(getattr(file_stream, "seek")):
                file_stream.seek(0)
        pdf_buffer.seek(0)

        patrimonios_agrupados = {}
        todos_patrimonios = set()
        current_item_material = None

        pdf_reader = PyPDF2.PdfReader(pdf_buffer)
        full_text = ""
        for page_num, page in enumerate(pdf_reader.pages):
            page_text = page.extract_text()
            if page_text:
                full_text += page_text + "\n"
            else:
                app.logger.warning(f"Não foi possível extrair texto da página {page_num + 1} do PDF.")

        lines = full_text.split("\n")
        item_material_pattern = re.compile(r"Item Material\s*:\s*([\w\d-]+)")
        patrimonio_pattern = re.compile(r"^\s*(\d{9})\b")

        for line in lines:
            line_stripped = line.strip()
            match_item_material = item_material_pattern.search(line_stripped)
            if match_item_material:
                current_item_material = match_item_material.group(1).strip()
                if current_item_material not in patrimonios_agrupados:
                    patrimonios_agrupados[current_item_material] = []
            
            if current_item_material:
                match_patrimonio = patrimonio_pattern.match(line_stripped)
                if match_patrimonio:
                    patrimonio = match_patrimonio.group(1)
                    todos_patrimonios.add(patrimonio)
                    if patrimonio not in patrimonios_agrupados[current_item_material]:
                        patrimonios_agrupados[current_item_material].append(patrimonio)
        
        if not todos_patrimonios:
            app.logger.warning("Nenhum patrimônio (formato de 9 dígitos) encontrado no PDF.")
        
        patrimonios_agrupados = {k: v for k, v in patrimonios_agrupados.items() if v}
        if not patrimonios_agrupados and todos_patrimonios:
             app.logger.warning("Patrimônios encontrados no PDF, mas nenhum pôde ser agrupado por Item Material. Verifique o formato do PDF.")
        elif not patrimonios_agrupados:
            app.logger.warning("Nenhum item material com patrimônios associados encontrado no PDF.")

        return patrimonios_agrupados, todos_patrimonios
        
    except Exception as e:
        app.logger.error(f"Erro ao processar PDF: {str(e)}", exc_info=True)
        raise ValueError(f"Erro ao processar o arquivo PDF: {str(e)}. Verifique se o arquivo é um PDF válido e não está corrompido.")

def extrair_patrimonios_excel(file_stream):
    patrimonios = set()
    original_filename = getattr(file_stream, "filename", "arquivo_excel")
    file_ext = os.path.splitext(original_filename)[1].lower()
    app.logger.debug(f"Iniciando extração de Excel para: {original_filename}, extensão: {file_ext}")

    excel_buffer = io.BytesIO()
    try:
        content = file_stream.read()
        excel_buffer.write(content)
        excel_buffer.seek(0)
        if hasattr(file_stream, "seek") and callable(getattr(file_stream, "seek")):
            try:
                file_stream.seek(0)
            except Exception as e_seek:
                app.logger.debug(f"Não foi possível fazer seek no file_stream original: {e_seek}")
    except Exception as e_buffer:
        app.logger.error(f"Erro ao preparar buffer para Excel \'{original_filename}\': {str(e_buffer)}", exc_info=True)
        raise ValueError(f"Erro ao ler dados do arquivo Excel/ODS \'{original_filename}\'.")

    try:
        col_idx_to_read = -1

        if file_ext in (".xlsx", ".xls"):
            app.logger.debug(f"Processando como .xlsx/.xls: {original_filename}")
            workbook = openpyxl.load_workbook(excel_buffer, data_only=True)
            sheet = workbook.active
            
            headers = []
            if sheet.max_row >= 1:
                headers = [str(cell.value) if cell.value is not None else "" for cell in sheet[1]]
            app.logger.info(f"Colunas encontradas (openpyxl) em \'{original_filename}\': {headers}")

            possible_names = ["patrimonio", "patrimônio", "numero de patrimonio", "número de patrimônio", "ativo", "plaqueta"]
            if headers:
                for i, header in enumerate(headers):
                    if any(name in header.lower() for name in possible_names):
                        col_idx_to_read = i
                        app.logger.info(f"Coluna de patrimônio encontrada pelo nome \'{header}\' no índice {i} em \'{original_filename}\'.")
                        break
            
            if col_idx_to_read == -1:
                if sheet.max_column >= 2:
                    col_idx_to_read = 1 
                    app.logger.info(f"Nenhuma coluna de patrimônio nomeada encontrada em \'{original_filename}\', usando a segunda coluna (índice 1).")
                elif sheet.max_column == 1:
                    col_idx_to_read = 0 
                    app.logger.info(f"Nenhuma coluna de patrimônio nomeada encontrada em \'{original_filename}\' e há apenas uma coluna, usando a primeira (índice 0).")
                else:
                    app.logger.warning(f"A planilha Excel \'{original_filename}\' parece estar vazia ou não contém colunas legíveis.")
            
            if col_idx_to_read != -1:
                app.logger.info(f"Usando coluna do Excel (índice {col_idx_to_read}) para patrimônios em \'{original_filename}\'.")
                start_row = 2 if headers and sheet.max_row > 1 else 1 # Pular cabeçalho se existir e houver dados
                for row_cells in sheet.iter_rows(min_row=start_row, values_only=True):
                    if row_cells and col_idx_to_read < len(row_cells):
                        cell_value = row_cells[col_idx_to_read]
                        if cell_value is not None:
                            s_value = str(cell_value).replace(".0", "").strip()
                            if re.fullmatch(r"\d{9}", s_value):
                                patrimonios.add(s_value)
        elif file_ext == ".ods":
            app.logger.debug(f"Processando como .ods: {original_filename}")
            doc = load_ods(excel_buffer)
            all_tables = doc.getElementsByType(Table)
            if not all_tables:
                 raise ValueError(f"Nenhuma tabela (planilha) encontrada no arquivo ODS: \'{original_filename}\'.")
            sheet = all_tables[0]
            
            rows = sheet.getElementsByType(TableRow)
            if not rows:
                app.logger.warning(f"Nenhuma linha encontrada na tabela do arquivo ODS \'{original_filename}\'.")
            else:
                headers = []
                if rows:
                    first_row_cells = rows[0].getElementsByType(TableCell)
                    for cell in first_row_cells:
                        text_content_elements = cell.getElementsByType(P)
                        text_content = "".join(str(te.firstChild) for te in text_content_elements if te.firstChild).strip()
                        headers.append(text_content)
                app.logger.info(f"Colunas encontradas (odfpy) em \'{original_filename}\': {headers}")

                possible_names = ["patrimonio", "patrimônio", "numero de patrimonio", "número de patrimônio", "ativo", "plaqueta"]
                if headers:
                    for i, header in enumerate(headers):
                        if any(name in header.lower() for name in possible_names):
                            col_idx_to_read = i
                            app.logger.info(f"Coluna de patrimônio encontrada pelo nome \'{header}\' no índice {i} em \'{original_filename}\'.")
                            break
                
                if col_idx_to_read == -1:
                    if len(headers) >= 2:
                        col_idx_to_read = 1
                        app.logger.info(f"Nenhuma coluna de patrimônio nomeada encontrada no ODS \'{original_filename}\', usando a segunda coluna (índice 1).")
                    elif len(headers) == 1:
                        col_idx_to_read = 0
                        app.logger.info(f"Nenhuma coluna de patrimônio nomeada encontrada no ODS \'{original_filename}\' e há apenas uma coluna, usando a primeira (índice 0).")
                    else:
                        app.logger.warning(f"A planilha ODS \'{original_filename}\' parece estar vazia ou não contém colunas legíveis.")

                if col_idx_to_read != -1:
                    app.logger.info(f"Usando coluna do ODS (índice {col_idx_to_read}) para patrimônios em \'{original_filename}\'.")
                    start_row_index = 1 if headers and len(rows) > 1 else 0
                    for row_idx in range(start_row_index, len(rows)):
                        row_element = rows[row_idx]
                        cells = row_element.getElementsByType(TableCell)
                        if col_idx_to_read < len(cells):
                            cell_obj = cells[col_idx_to_read]
                            s_value = ""
                            for p_element in cell_obj.getElementsByType(P):
                                if p_element.firstChild:
                                    s_value += str(p_element.firstChild)
                            s_value = s_value.strip().replace(".0", "")
                            if re.fullmatch(r"\d{9}", s_value):
                                patrimonios.add(s_value)
        else:
            raise ValueError(f"Formato de arquivo não suportado: {original_filename}. Use .xlsx, .xls ou .ods.")

        if not patrimonios:
            app.logger.warning(f"Nenhum patrimônio (formato de 9 dígitos) encontrado no arquivo \'{original_filename}\' na coluna processada.")
            
        return patrimonios
    
    except ValueError as ve:
        app.logger.error(f"Erro de valor ao processar arquivo \'{original_filename}\': {str(ve)}")
        raise
    except Exception as e:
        app.logger.error(f"Erro inesperado ao processar arquivo \'{original_filename}\': {str(e)}", exc_info=True)
        raise ValueError(f"Erro ao processar o arquivo Excel/ODS \'{original_filename}\': {str(e)}. Verifique o formato e conteúdo do arquivo.")

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        try:
            if "excel" not in request.files or "pdf" not in request.files:
                raise ValueError("Ambos os arquivos (Excel e PDF) são obrigatórios.")
                
            excel_file = request.files["excel"]
            pdf_file = request.files["pdf"]
            
            if not excel_file or excel_file.filename == "" or not pdf_file or pdf_file.filename == "":
                raise ValueError("Nenhum arquivo selecionado ou nome de arquivo inválido. Por favor, selecione ambos os arquivos.")
            
            excel_filename = secure_filename(excel_file.filename)
            pdf_filename = secure_filename(pdf_file.filename)
            
            app.logger.info(f"Processando Excel: {excel_filename} e PDF: {pdf_filename}")

            patrimonios_excel = extrair_patrimonios_excel(excel_file)
            patrimonios_agrupados_pdf, todos_patrimonios_pdf = extrair_dados_pdf_refatorado(pdf_file)
            
            app.logger.info(f"Patrimônios Excel ({len(patrimonios_excel)}): {list(patrimonios_excel)[:5]}...")
            app.logger.info(f"Todos Patrimônios PDF ({len(todos_patrimonios_pdf)}): {list(todos_patrimonios_pdf)[:5]}...")
            app.logger.info(f"Patrimônios Agrupados PDF (amostra): { {k: v[:2] for k, v in list(patrimonios_agrupados_pdf.items())[:2]} }...")

            patrimonios_ambos = patrimonios_excel & todos_patrimonios_pdf
            somente_excel = patrimonios_excel - todos_patrimonios_pdf
            somente_pdf = todos_patrimonios_pdf - patrimonios_excel
            
            resultados_json = {
                "total_excel": len(patrimonios_excel),
                "total_pdf": len(todos_patrimonios_pdf),
                "total_ambos": len(patrimonios_ambos),
                "total_somente_excel": len(somente_excel),
                "total_somente_pdf": len(somente_pdf),
                "excel_filename": excel_filename,
                "pdf_filename": pdf_filename,
                "patrimonios_agrupados_pdf": [],
                "exemplos_excel": sorted(list(somente_excel))[:4] if somente_excel else [],
                "mostrar_exemplos": len(somente_excel) > 0
            }

            patrimonios_agrupados_json_view = []
            if patrimonios_agrupados_pdf:
                for item, pats in sorted(patrimonios_agrupados_pdf.items()):
                    if pats:
                         patrimonios_agrupados_json_view.append(f"Item {item}: {', '.join(sorted(list(set(pats))))}")
            if not patrimonios_agrupados_json_view:
                patrimonios_agrupados_json_view = ["Nenhum patrimônio agrupado encontrado no PDF."]
            resultados_json["patrimonios_agrupados_pdf"] = patrimonios_agrupados_json_view
            
            output_excel = io.BytesIO()
            wb = openpyxl.Workbook()
            
            ws_resumo = wb.active
            ws_resumo.title = "Resumo Geral"
            ws_resumo.append(["Descrição", "Quantidade"])
            resumo_data_excel = [
                (f"Total de patrimônios no Excel ({excel_filename})", resultados_json["total_excel"]),
                (f"Total de patrimônios no PDF ({pdf_filename} - geral)", resultados_json["total_pdf"]),
                ("Patrimônios encontrados em ambas as fontes", resultados_json["total_ambos"]),
                (f"Patrimônios encontrados apenas no Excel ({excel_filename})", resultados_json["total_somente_excel"]),
                (f"Patrimônios encontrados apenas no PDF ({pdf_filename} - geral)", resultados_json["total_somente_pdf"])
            ]
            for row_data in resumo_data_excel:
                ws_resumo.append(row_data)

            ws_agrupados = wb.create_sheet(title="Patrimonios por Item (PDF)")
            ws_agrupados.append(["Item Material (do PDF)", "Patrimônio Associado"])
            if patrimonios_agrupados_pdf:
                for item_material, pats in sorted(patrimonios_agrupados_pdf.items()):
                    if pats:
                        for pat in sorted(list(set(pats))):
                            ws_agrupados.append([item_material, pat])
            else:
                ws_agrupados.append(["Nenhum patrimônio agrupado por item material encontrado no PDF.", ""])

            ws_ambos = wb.create_sheet(title="Patrimonios em Ambos")
            ws_ambos.append(["Patrimônios em Ambos"])
            if patrimonios_ambos:
                for pat in sorted(list(patrimonios_ambos)):
                    ws_ambos.append([pat])
            else:
                ws_ambos.append(["Nenhum patrimônio em comum"])

            ws_so_excel = wb.create_sheet(title="Apenas no Excel")
            ws_so_excel.append(["Somente no Excel"])
            if somente_excel:
                for pat in sorted(list(somente_excel)):
                    ws_so_excel.append([pat])
            else:
                ws_so_excel.append(["Nenhum patrimônio exclusivo do Excel"])

            ws_so_pdf = wb.create_sheet(title="Apenas no PDF (Geral)")
            ws_so_pdf.append(["Somente no PDF (Geral)"])
            if somente_pdf:
                for pat in sorted(list(somente_pdf)):
                    ws_so_pdf.append([pat])
            else:
                ws_so_pdf.append(["Nenhum patrimônio exclusivo do PDF (Geral)"])
            
            wb.save(output_excel)
            output_excel.seek(0)
            
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify({
                    "success": True,
                    "resumo": resultados_json
                })
            
            return send_file(
                output_excel,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                download_name="resultado_comparacao_patrimonios.xlsx",
                as_attachment=True
            )
            
        except ValueError as ve:
            app.logger.warning(f"Erro de Validação/Processamento na rota: {str(ve)}")
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify({"success": False, "erro": str(ve)})
            return render_template("index.html", erro=str(ve))

        except Exception as e:
            app.logger.error(f"Erro Inesperado na rota: {str(e)}", exc_info=True)
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify({"success": False, "erro": "Ocorreu um erro inesperado no servidor."})
            return render_template("index.html", erro="Ocorreu um erro inesperado no servidor. Tente novamente.")
    
    return render_template("index.html")

if __name__ == '__main__':
    if not app.logger.handlers or not any(isinstance(h, logging.StreamHandler) for h in app.logger.handlers):
        app.logger.handlers.clear()
        stream_handler = logging.StreamHandler(sys.stdout)
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        stream_handler.setFormatter(formatter)
        app.logger.addHandler(stream_handler)
        app.logger.setLevel(logging.DEBUG)
    
    app.logger.info("Aplicação Flask Validador de Patrimônios iniciando em modo de debug local...")
    app.run(debug=True, host="0.0.0.0", port=5000)

